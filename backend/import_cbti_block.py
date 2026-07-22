"""One-shot import of a COMPLETED CBT-I block from a VA CBT-I Sleep Diary
Calculator export (Rachel Manber, 2010; VA CBT-I Training Program).

NOT a migration. Loads one cbti_block, its cbti_prescriptions (derived from
`Rx Bedtime` change-points), and per-night sleep-diary fields onto daily_records.

The workbook holds dated sleep, alcohol counts, and free-text — BOTH repos are
public. This script takes the path as an argument and reads it from wherever it
lives (never copied into the tree); the workbook and any fixture derived from it
are git-ignored (cbti_user_data_*.xlsx). Never commit the file or real rows.

Reconciliation (Gate 4): per-night SE is recomputed independently
(TST / TIB, TIB = Out-of-Bed - Time-Tried-to-Sleep with conditional midnight
wrap) and checked against the sheet's own `Sleep Efficiency` to +/-0.001. A
negative control perturbs one night by 0.01 and asserts the checker flags
exactly that night — proving the comparator can see a mismatch before we trust
that it sees none. The load ABORTS if any real night mismatches.

Field mapping (daily_records <- sheet column):
    lights_out         <- Time Tried to Sleep   (SE window opens here; HH:MM)
    sleep_latency_min  <- Latency Amount         (minutes)
    waso_min           <- Time Awake             (minutes; == WASO col where present)
    night_wakings_n    <- Wakeup Count
    final_wake         <- Wakeup Time            (HH:MM)
    out_of_bed         <- Out of Bed Time        (HH:MM)
    naps_min           <- Nap Duration           (minutes; stored on the diary's OWN
                          row-date, faithful to source — NOT shifted. See NAP NOTE.)
    diary_se_pct       <- Sleep Efficiency * 100 (percent, matching samsung sleep_efficiency_pct)
    diary_tst_min      <- Total Sleep Time       (minutes)
    alcohol_units      <- Alcohol                ('No'->0, ''->None, int units).
                          Beyond the 9 diary fields; imported because the engine's
                          exclusion (alcohol_units > 0) needs it and rows are created now.

NAP NOTE (silent-when-wrong): the engine reads naps for the night terminating
on wake-date W from daily_records.date = W-1. This import stores each nap on the
same row-date the source recorded it against; it does NOT shift by a day. Whether
the VA diary's nap item refers to naps preceding vs following the recorded night
must be confirmed against the instrument before the engine relies on the date-1
read. Only 2 nights carry naps in this block.

Usage:
    python backend/import_cbti_block.py --xlsx "<path>" --user-id 1            # dry-run (default)
    python backend/import_cbti_block.py --xlsx "<path>" --user-id 1 --apply    # write to DB (DATABASE_URL)
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys

import openpyxl

import models
from database import SessionLocal

SHEET = "Sleep Diaries For Import"
WAKE_ANCHOR = "05:00"
SE_TOL = 0.001


# ── parsing helpers ─────────────────────────────────────────────────────────────

def _tmin(v):
    """A datetime.time -> minutes since midnight, else None."""
    return v.hour * 60 + v.minute if isinstance(v, dt.time) else None


def _hhmm(v):
    return f"{v.hour:02d}:{v.minute:02d}" if isinstance(v, dt.time) else None


def _tib(tts_min, oob_min):
    """Total time in bed = Out-of-Bed - Time-Tried-to-Sleep, wrapping past midnight
    only when Out-of-Bed clock time is earlier than Time-Tried-to-Sleep."""
    if tts_min is None or oob_min is None:
        return None
    return (oob_min - tts_min) if oob_min >= tts_min else (oob_min + 1440 - tts_min)


def _parse_alcohol(v):
    if v is None or v == "" or v == "No":
        return 0 if v == "No" else None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        m = re.match(r"\d+", v.strip())
        return int(m.group()) if m else None
    return None


def _window_minutes(lights_out_hhmm: str, wake_hhmm: str) -> int:
    lh, lm = map(int, lights_out_hhmm.split(":"))
    wh, wm = map(int, wake_hhmm.split(":"))
    return ((wh * 60 + wm) + 24 * 60) - (lh * 60 + lm)


# ── workbook -> structured nights + prescriptions ───────────────────────────────

def parse_workbook(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet {SHEET!r} not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    def col(r, name):
        return ws.cell(r, hdr[name]).value

    rows = [r for r in range(2, ws.max_row + 1) if isinstance(col(r, "Diary Date"), dt.datetime)]
    nights = []
    for r in rows:
        tts = _tmin(col(r, "Time Tried to Sleep"))
        oob = _tmin(col(r, "Out of Bed Time"))
        tst = _tmin(col(r, "Total Sleep Time"))
        tib = _tib(tts, oob)
        nights.append({
            "date": col(r, "Diary Date").date(),
            "lights_out": _hhmm(col(r, "Time Tried to Sleep")),
            "sleep_latency_min": _tmin(col(r, "Latency Amount")),
            "waso_min": _tmin(col(r, "Time Awake")),
            "night_wakings_n": col(r, "Wakeup Count"),
            "final_wake": _hhmm(col(r, "Wakeup Time")),
            "out_of_bed": _hhmm(col(r, "Out of Bed Time")),
            "naps_min": _tmin(col(r, "Nap Duration")) or 0,
            "diary_tst_min": tst,
            "alcohol_units": _parse_alcohol(col(r, "Alcohol")),
            "rx_lights_out": _hhmm(col(r, "Rx Bedtime")),
            "rx_wake": _hhmm(col(r, "Rx Wake Time")),
            "_tib": tib,
            "_ref_se": col(r, "Sleep Efficiency"),           # sheet fraction, for reconcile
            "recomp_se": (tst / tib) if (tst and tib) else None,
        })
    nights.sort(key=lambda n: n["date"])
    for n in nights:
        # frozen diary SE stored as PERCENT (samsung sleep_efficiency_pct convention)
        n["diary_se_pct"] = round(n["_ref_se"] * 100, 2) if n["_ref_se"] is not None else None

    # prescriptions = Rx Bedtime change-points, chronological
    prescriptions = []
    last = None
    for n in nights:
        rx = n["rx_lights_out"]
        if rx is not None and rx != last:
            prescriptions.append({"effective_from": n["date"], "lights_out": rx,
                                  "wake": n["rx_wake"] or WAKE_ANCHOR})
            last = rx
    for i, p in enumerate(prescriptions):
        p["window_minutes"] = _window_minutes(p["lights_out"], p["wake"])
        if i == 0:
            p["decision"] = "adopt"
        else:
            prev = prescriptions[i - 1]["window_minutes"]
            p["decision"] = ("extend" if p["window_minutes"] > prev
                             else "compress" if p["window_minutes"] < prev else "hold")
    return nights, prescriptions


# ── reconciliation + negative control ───────────────────────────────────────────

def reconcile(nights, tol=SE_TOL):
    """Return list of (date, delta) for nights whose independent SE recompute
    disagrees with the sheet's SE beyond tol. Comparison is on the fraction scale."""
    mm = []
    for n in nights:
        rec, ref = n["recomp_se"], n["_ref_se"]
        if rec is None or ref is None:
            mm.append((n["date"], "missing"))
        elif abs(rec - ref) > tol:
            mm.append((n["date"], round(rec - ref, 4)))
    return mm


def reconcile_with_control(nights):
    real = reconcile(nights)
    # negative control: perturb one night's reference SE by +0.01, expect exactly it flagged
    import copy
    idx = len(nights) // 2
    pert = copy.deepcopy(nights)
    pert[idx]["_ref_se"] = (pert[idx]["_ref_se"] or 0) + 0.01
    ctrl = reconcile(pert)
    ctrl_new = [m for m in ctrl if m not in real]
    control_ok = (len(ctrl_new) == 1 and ctrl_new[0][0] == pert[idx]["date"])
    worst = max((abs(n["recomp_se"] - n["_ref_se"]) for n in nights
                 if n["recomp_se"] is not None and n["_ref_se"] is not None), default=None)
    return {"real": real, "control_ok": control_ok, "control_date": pert[idx]["date"],
            "control_flagged": ctrl_new, "worst_residual": worst}


# ── load ────────────────────────────────────────────────────────────────────────

def load(db, user_id, nights, prescriptions, close_date, apply):
    existing = (db.query(models.CBTIBlock)
                .filter_by(user_id=user_id, opened_on=nights[0]["date"]).first())
    if existing:
        sys.exit(f"ABORT: a cbti_block already exists for user {user_id} opened_on "
                 f"{nights[0]['date']} (id={existing.id}). Refusing to double-import.")

    # exit summary over the final prescription's nights (descriptive aggregate)
    final_from = prescriptions[-1]["effective_from"]
    final_nights = [n for n in nights if final_from <= n["date"] <= close_date]
    exit_tst = round(sum(n["diary_tst_min"] for n in final_nights) / len(final_nights)) if final_nights else None
    exit_se = round(sum(n["diary_se_pct"] for n in final_nights) / len(final_nights), 1) if final_nights else None

    if not apply:
        print(f"[dry-run] would create block user={user_id} opened={nights[0]['date']} "
              f"closed={close_date} exit_tst={exit_tst}m exit_se={exit_se}%")
        print(f"[dry-run] would create {len(prescriptions)} prescriptions, {len(nights)} daily_records")
        return None

    block = models.CBTIBlock(
        user_id=user_id, opened_on=nights[0]["date"], closed_on=close_date,
        wake_anchor=WAKE_ANCHOR,
        open_reason="Third CBT-I block (two prior completions). Imported from VA CBT-I "
                    "Sleep Diary Calculator export.",
        close_reason="TST plateau at 7h38 window with SE held >=85% (#104 exit condition); "
                     "sleep need ~7h30 (#101).",
        exit_tst_min=exit_tst, exit_se_pct=exit_se,
        notes="Source: VA CBT-I Sleep Diary Calculator (Manber 2010). 53 nights, 3 gaps.",
    )
    db.add(block)
    db.flush()

    rx_rows = []
    for i, p in enumerate(prescriptions):
        nxt = prescriptions[i + 1]["effective_from"] if i + 1 < len(prescriptions) else None
        eff_to = (nxt - dt.timedelta(days=1)) if nxt else close_date
        rx = models.CBTIPrescription(
            block_id=block.id, effective_from=p["effective_from"], effective_to=eff_to,
            prescribed_lights_out=p["lights_out"], wake_anchor=p["wake"],
            window_minutes=p["window_minutes"], decision=p["decision"],
            rationale="Historical import: window derived from Rx Bedtime change-point; "
                      "basis_* left null (engine computes on replay, Gate 5).",
        )
        db.add(rx)
        db.flush()
        rx_rows.append(rx)
    # chain supersession (append-only whitelisted updates)
    for i in range(len(rx_rows) - 1):
        rx_rows[i].superseded_by = rx_rows[i + 1].id

    for n in nights:
        db.add(models.DailyRecord(
            user_id=user_id, date=n["date"],
            lights_out=n["lights_out"], sleep_latency_min=n["sleep_latency_min"],
            waso_min=n["waso_min"], night_wakings_n=n["night_wakings_n"],
            final_wake=n["final_wake"], out_of_bed=n["out_of_bed"],
            naps_min=n["naps_min"], diary_se_pct=n["diary_se_pct"],
            diary_tst_min=n["diary_tst_min"], alcohol_units=n["alcohol_units"],
        ))
    db.commit()
    print(f"COMMITTED: block id={block.id}, {len(rx_rows)} prescriptions, {len(nights)} daily_records.")
    return block.id


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="path to the VA CBT-I diary workbook")
    ap.add_argument("--user-id", type=int, required=True)
    ap.add_argument("--apply", action="store_true", help="write to DB (default: dry-run)")
    args = ap.parse_args()

    nights, prescriptions = parse_workbook(args.xlsx)
    close_date = dt.date(2026, 5, 11)

    print(f"parsed: {len(nights)} nights {nights[0]['date']}..{nights[-1]['date']}, "
          f"{len(prescriptions)} prescriptions")
    for p in prescriptions:
        print(f"  rx {p['effective_from']} lo={p['lights_out']} win={p['window_minutes']}m "
              f"{p['decision']}")

    rc = reconcile_with_control(nights)
    print(f"reconcile: real_mismatches={len(rc['real'])} worst_residual={rc['worst_residual']:.6f} "
          f"control_ok={rc['control_ok']} (perturbed {rc['control_date']}, "
          f"flagged {[str(d) for d, _ in rc['control_flagged']]})")
    if not rc["control_ok"]:
        sys.exit("ABORT: negative control failed — the SE checker did not localize an injected "
                 "mismatch; do not trust a clean pass.")
    if rc["real"]:
        sys.exit(f"ABORT: {len(rc['real'])} real SE mismatch(es) > {SE_TOL}: {rc['real']}")

    db = SessionLocal()
    try:
        load(db, args.user_id, nights, prescriptions, close_date, args.apply)
    finally:
        db.close()


if __name__ == "__main__":
    main()
